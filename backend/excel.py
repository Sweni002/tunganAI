from flask import Flask, send_file, request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
from datetime import datetime, time, date

app = Flask(__name__)


def minutes_to_hms(minutes):
    try:
        minutes = int(minutes)  # ← conversion sécurisée
    except (ValueError, TypeError):
        minutes = 0

    if minutes == 0:
        return "00:00:00"

    heures = minutes // 60
    reste = minutes % 60
    return f"{heures:02d}:{reste:02d}:00"


def creer_fiche_assiduite(
    nom_service,
    mois,
    annee,
    sigle_service_adresse,
    date_debut,
    date_fin,
    structure_divisions,
    types_absences,  # liste de dict { "idtype": 1, "abbreviation": "RAM", "nom": "Repos annuel" }
):
    """
    Génère le fichier Excel avec les paramètres dynamiques et les divisions personnalisées.
    Colonnes D-F = nombre de retards, volume de retard (mn), nombre de JA non justifiés
    Colonnes G+ = absences par type
    """
    os.makedirs("excel_output", exist_ok=True)
    filepath = f"excel_output/fiche_assiduite_{mois}_{annee}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = mois.upper()

    # --- LOGO ---
    ws.merge_cells("B1:F3")
    if os.path.exists("logo.png"):
        logo = Image("logo.png")
        logo.width = 200
        logo.height = 120
        ws.add_image(logo, "F1")

    # --- TEXTE INSTITUTIONNEL ---
    textes = [
        "SECRETARIAT GENERAL",
        "DIRECTION GENERAL DU BUDGET",
        "ET DES FINANCES",
        "DIRECTION DE LA SOLDE ET DES PENSIONS",
        nom_service.upper(),
    ]
    for i, texte in enumerate(textes):
        cell = ws.cell(row=9 + i, column=2)
        cell.value = texte
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["B"].width = 45
    for r in range(9, 15):
        ws.row_dimensions[r].height = 22

    # --- TITRE ---
    ws.merge_cells("B16:F16")
    ws["B16"] = f"FICHE DE CONTRÔLE D’ASSIDUITÉ : {mois.upper()} {annee}"
    ws["B16"].font = Font(bold=True, size=14)
    ws["B16"].alignment = Alignment(horizontal="center")

    # --- OBJET ET PÉRIODES ---
    ws["A17"] = (
        f"Objet: Suivi Mensuel de l'absentéisme du personnel au sein du {sigle_service_adresse}"
    )
    ws["A17"].font = Font(bold=True, size=10)
    ws.merge_cells("G17:K17")
    ws["G17"] = f"Périodes: Du {date_debut} Au {date_fin}"
    ws["G17"].font = Font(bold=True, size=10)
    ws["G17"].alignment = Alignment(horizontal="center", vertical="center")

    # --- TABLEAU (Headers) ---
    header_fill = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers fixes
    ws.merge_cells("A20:A21")
    ws["A20"] = "N°"
    ws.merge_cells("B20:B21")
    ws["B20"] = "NOM ET PRENOMS"
    ws.merge_cells("C20:C21")
    ws["C20"] = "IM"

    # Section Non Valables
    ws.merge_cells("D20:F20")
    ws["D20"] = "Volume et Nombre de jrs d’absences non valables"
    for i, t in enumerate(
        [
            "Nombre de retard",
            "Volume de retard(en mn)",
            "Nombre de JA non justifié ou non valable",
        ]
    ):
        ws.cell(row=21, column=4 + i).value = t

    # Section Valables – Colonnes dynamiques selon les types
    col_start = 7
    nb_types = len(types_absences)
    col_end = col_start + nb_types - 1
    ws.merge_cells(start_row=20, start_column=col_start, end_row=20, end_column=col_end)
    ws.cell(row=20, column=col_start).value = (
        "Nombre de jrs d’absences justifiées ou valables(jrs)"
    )

    # Sous-headers = abréviations des types
    for i, t in enumerate(types_absences):
        ws.cell(row=21, column=col_start + i).value = t.get("nomtype", t.get("nom", ""))

    # Style Header
    for r in range(20, 22):
        for c in range(1, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # --- DONNÉES DYNAMIQUES PAR DIVISIONS ---
    current_row = 22
    division_fill = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    compteur_global = 1
    for div_name, staff_list in structure_divisions.items():
        # 1. Fusion à partir de B
        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=col_end,
        )

        # Définition des bordures spécifiques
        thin_side = Side(style="thin")
        no_side = Side(style=None)

        # 2. Boucle sur les colonnes
        for c in range(1, col_end + 1):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = division_fill
            cell.font = Font(bold=True, size=11)

            # --- LOGIQUE DES BORDURES POUR ENLEVER LA SÉPARATION ---
            if c == 1:  # Colonne A (N°)
                cell.border = Border(
                    left=thin_side, top=thin_side, bottom=thin_side, right=no_side
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 2:  # Colonne B (Début du groupe fusionné)
                cell.value = div_name
                # On enlève la bordure gauche pour "coller" à la colonne A
                cell.border = Border(
                    left=no_side, top=thin_side, bottom=thin_side, right=no_side
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c == col_end:  # Dernière colonne
                cell.border = Border(
                    left=no_side, top=thin_side, bottom=thin_side, right=thin_side
                )
            else:  # Colonnes intermédiaires du bloc fusionné
                cell.border = Border(top=thin_side, bottom=thin_side)

        ws.row_dimensions[current_row].height = 28

        current_row += 1

        # ... (reste du code pour les membres)

        # Lignes membres
        for person in staff_list:
            ws.cell(row=current_row, column=1).value = compteur_global
            ws.cell(row=current_row, column=2).value = person.get("nom", "")
            ws.cell(row=current_row, column=3).value = person.get("im", "")

            # --- Colonnes D-F = données dynamiques des pointages ---
            ws.cell(row=current_row, column=4).value = person.get("nombre_retards", 0)
            cell = ws.cell(row=current_row, column=5)
            cell.value = minutes_to_hms(person.get("volume_retards", 0))

            ws.cell(row=current_row, column=6).value = person.get(
                "nombre_ja_non_justifiees", 0
            )
            for col_idx in range(4, 7):
                ws.cell(row=current_row, column=col_idx).alignment = Alignment(
                    horizontal="right", vertical="center"
                )

            # Colonnes valables = absences_par_type
               # Colonnes valables = absences_par_type (corrigé pour liste → dict)
            absences_dict = {a["abbreviation"]: a["nombre"] for a in person.get("absences_par_type", [])}

            for i, t in enumerate(types_absences):
               col_idx = col_start + i
               value = absences_dict.get(t["abbreviation"], 0)
               ws.cell(row=current_row, column=col_idx).value = value
               ws.cell(row=current_row, column=col_idx).alignment = Alignment(
        horizontal="center", vertical="center"
    )


            # Bordures et alignement des 3 premières colonnes
            for col_idx in range(1, col_end + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                if col_idx <= 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            compteur_global += 1
            current_row += 1

    # Ajustements finaux des colonnes
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 12
    for col in range(4, col_end + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    wb.save(filepath)
    return filepath


def creer_fiche_presence(
    nom_service,
    date_jour,
    sigle_service_adresse,
    structure_divisions,
    logo_file="logo_temp.png",
):
    os.makedirs("excel_output", exist_ok=True)
    filepath = f"excel_output/fiche_presence_{date_jour.replace('/', '_')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PRESENCE"

    # --- LOGO ---
    if os.path.exists("logo.png"):
        logo = Image("logo.png")
        logo.width = 200
        logo.height = 120
        ws.add_image(logo, "G1")

    # --- EN-TETE ---
    textes = [
        "SECRETARIAT GENERAL",
        "DIRECTION GENERAL DU BUDGET",
        "ET DES FINANCES",
        "DIRECTION DE LA SOLDE ET DES PENSIONS",
        nom_service.upper(),
    ]
    for i, t in enumerate(textes):
          cell = ws.cell(row=7 + i, column=2, value=t)
          cell.font = Font(bold=True, size=12)
          cell.alignment = Alignment(horizontal="center", vertical="center" )  # ✅ Centrer horizontalement

    # --- TITRE ---
    ws.merge_cells("B13:M13")
    ws["B13"] = f"FICHE DE PRÉSENCE : {date_jour}"
    ws["B13"].font = Font(bold=True, size=14)
    ws["B13"].alignment = Alignment(horizontal="center")
    ws["B14"] = f"Objet:Fiche de presence au sein du {sigle_service_adresse}"

    # --- HEADERS ---
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Configuration des colonnes
    ws.merge_cells("A17:A18")
    ws["A17"] = "N°"
    ws.merge_cells("B17:B18")
    ws["B17"] = "NOM ET PRENOMS"
    ws.merge_cells("C17:C18")
    ws["C17"] = "IM"
    ws.merge_cells("D17:G17")
    ws["D17"] = "MATIN"
    ws.merge_cells("H17:K17")
    ws["H17"] = "APRES-MIDI"
    ws.merge_cells("L17:M17")
    ws["L17"] = "JUSTIFICATIFS"

    sub = [
        "Entrée",
        "Sortie",
        "Retard",
        "Absent",
        "Entrée",
        "Sortie",
        "Retard",
        "Absent",
        "Matin",
        "A-M",
    ]
    for i, s in enumerate(sub):
        ws.cell(row=18, column=4 + i, value=s)

    for r in range(17, 19):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- DATA ---
    row = 19
    idx = 1
    division_fill = PatternFill("solid", fgColor="BDD7EE") 

    for div, staff in structure_divisions.items():
        # 1. Ajustement de la hauteur de la ligne à 28
        ws.row_dimensions[row].height = 28

        # 2. Fusion et contenu de la ligne de division
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        cell_div = ws.cell(row=row, column=1, value=div)

        # 3. Application du style (Bleu, Gras, Alignement vertical centré)
        cell_div.font = Font(bold=True, size=11)
        cell_div.fill = division_fill
        cell_div.alignment = Alignment(horizontal="left", vertical="center")

        # 4. Application des bordures sur toute la largeur de la ligne fusionnée
        for c in range(1, 14):
            ws.cell(row=row, column=c).border = border

        row += 1

        for p in staff:
            print(f"Traitement de l'employé : {p['nom']} (IM: {p['im']})")  # LOG
            if "heure_entree_unique" in p:
                print(
                    f"heure_entree_unique trouvée : {p['heure_entree_unique']} {p['justif_matin']}"
                )  # LOG
            else:
                print("Pas d'heure_entree_unique pour cet employé")  # LOG

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=p["nom"])
            ws.cell(row=row, column=3, value=p["im"])

            # Si agent de surface avec heure_entree_unique
            if "heure_entree_unique" in p and p["heure_entree_unique"]:
                h_entree = datetime.strptime(p["heure_entree_unique"], "%H:%M").time()
                # Entrée Matin ou Après-midi selon l'heure
                if h_entree < time(13, 0):
                    ws.cell(row=row, column=4, value=p.get("heure_entree_unique", "00:00"))  # Entrée Matin
     
                    ws.cell(row=row, column=5, value=p.get("heure_sortie_unique") or "00:00")  # Sortie Matin
                    ws.cell(row=row, column=8, value="---")
                    ws.cell(row=row, column=9, value="---")
  
                else:
                     ws.cell(row=row, column=4, value="---")
                     ws.cell(row=row, column=5, value="---")
                     ws.cell(row=row, column=8,value=p.get("heure_entree_unique", "00:00"))  # Entrée Après-midi
                     ws.cell(row=row, column=9,value=p.get("heure_sortie_unique") or "00:00")  # Sortie Après-midi
  # Entrée Après-midi

                # Colonnes inutiles pour agent surface
               
                ws.cell(row=row, column=6, value="---")  # Retard Matin
                ws.cell(row=row, column=7, value=p["absence_unique"])  # Absent Matin
                ws.cell(row=row, column=10, value="---")  # Retard Après-midi
                ws.cell(row=row, column=11,value=p["absence_unique"])  # Absent Après-midi
                ws.cell(row=row, column=12, value=p["justif_matin"])  # Justificatif Matin
                ws.cell(row=row, column=13, value=p["justif_matin"])  # Justificatif Après-midi

            else:
                # Logique classique pour les autres personnels
                ws.cell(row=row, column=4, value=p["matin_entree"])
                ws.cell(row=row, column=5, value=p["matin_sortie"])
                ws.cell(row=row, column=6, value=p["matin_retard"])
                ws.cell(row=row, column=7, value=p["matin_absent"])
                ws.cell(row=row, column=8, value=p["soir_entree"])
                ws.cell(row=row, column=9, value=p["soir_sortie"])
                ws.cell(row=row, column=10, value=p["soir_retard"])
                ws.cell(row=row, column=11, value=p["soir_absent"])
                ws.cell(row=row, column=12, value=p["justif_matin"])
                ws.cell(row=row, column=13, value=p["justif_soir"])

            # Bordures et alignement
            for c in range(1, 14):
                cell = ws.cell(row=row, column=c)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if c == 2 else "center", vertical="center"
                )

            row += 1
            idx += 1

    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["L"].width = 40  # Largeur augmentée pour le type d'absence
    ws.column_dimensions["M"].width = 40  # Largeur augmentée pour le type d'absence
    wb.save(filepath)
    return filepath


def creer_fiche_presence_periode(
    nom_service,
    periode_str,
    sigle_service_adresse,
    structure_divisions,
    logo_file="logo_temp.png",
):
    os.makedirs("excel_output", exist_ok=True)
    filepath = f"excel_output/fiche_presence_periode.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PRESENCE PERIODE"

    # --- LOGO ---
    if os.path.exists("logo.png"):
        logo = Image("logo.png")
        logo.width = 200
        logo.height = 140
        ws.add_image(logo, "G1")

    # --- EN-TETE ---
    textes = [
        "SECRETARIAT GENERAL",
        "DIRECTION REGIONALE DU BUDGET",
        "DIRECTION DE LA SOLDE",
        nom_service.upper(),
    ]
    for i, t in enumerate(textes):
        ws.cell(row=7 + i, column=2, value=t).font = Font(bold=True, size=12)

    # --- TITRE ---
    # Fusion sur 14 colonnes (A à N)
    ws.merge_cells("B13:N13")
    ws["B13"] = f"FICHE DE PRÉSENCE : {periode_str}"
    ws["B13"].font = Font(bold=True, size=14)
    ws["B13"].alignment = Alignment(horizontal="center")
    ws["B14"] = f"Objet: Fiche de présence au sein du {sigle_service_adresse}"

    # --- HEADERS ---
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Configuration des colonnes (Décalage pour inclure la DATE)
    ws.merge_cells("A17:A18")
    ws["A17"] = "DATE"
    ws.merge_cells("B17:B18")
    ws["B17"] = "N°"
    ws.merge_cells("C17:C18")
    ws["C17"] = "NOM ET PRENOMS"
    ws.merge_cells("D17:D18")
    ws["D17"] = "IM"
    ws.merge_cells("E17:H17")
    ws["E17"] = "MATIN"
    ws.merge_cells("I17:L17")
    ws["I17"] = "APRES-MIDI"
    ws.merge_cells("M17:N17")
    ws["M17"] = "JUSTIFICATIFS"

    sub_headers = [
        "Entrée",
        "Sortie",
        "Retard",
        "Absent",  # Matin (E-H)
        "Entrée",
        "Sortie",
        "Retard",
        "Absent",  # Soir (I-L)
        "Matin",
        "A-M",  # Justif (M-N)
    ]
    for i, s in enumerate(sub_headers):
        ws.cell(row=18, column=5 + i, value=s)

    # Style des en-têtes
    for r in range(17, 19):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # --- DATA ---
    row = 19
    idx = 1
    division_fill = PatternFill("solid", fgColor="BDD7EE")

    for div, entries in structure_divisions.items():
        # Ligne Division : Hauteur 28, Bleu, Gras
        ws.row_dimensions[row].height = 28
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        cell_div = ws.cell(row=row, column=1, value=div)
        cell_div.font = Font(bold=True, size=11)
        cell_div.fill = division_fill
        cell_div.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(1, 15):
            ws.cell(row=row, column=c).border = border
        row += 1

        # Lignes de pointage
        for p in entries:
            if "matin_absent_unique" in p:
                print(
                    f"matin_absent_unique trouvée : {p['matin_absent_unique']}"
                )  # LOG
            else:
                print("Pas d'matin_absent_unique pour cet employé")  # LOG

            ws.cell(row=row, column=1, value=p["date"])
            ws.cell(row=row, column=2, value=idx)
            ws.cell(row=row, column=3, value=p["nom"])
            ws.cell(row=row, column=4, value=p["im"])

            if "heure_entree_unique" in p and p["heure_entree_unique"]:
                h_entree = datetime.strptime(p["heure_entree_unique"], "%H:%M").time()
                # Entrée Matin ou Après-midi selon l'heure
                if h_entree < time(13, 0):
                    ws.cell(row=row, column=5, value=p["heure_entree_unique"] or '00:00')  # Entrée Matin
     
                    ws.cell(row=row, column=6, value=p["heure_sortie_unique"]or '00:00')  # Sortie Matin
                    ws.cell(row=row, column=9, value="---")
                    ws.cell(row=row, column=10, value="---")
  
                   
                else:
                     ws.cell(row=row, column=5, value="---")
                     ws.cell(row=row, column=6, value="---")
                     ws.cell(row=row, column=9, value=p["heure_entree_unique"])  # Entrée Après-midi
                     ws.cell(row=row, column=10, value=p["heure_sortie_unique"])  # Sortie Après-midi
  # Entrée Après-midi

                
                ws.cell(row=row, column=8, value=p["absence_unique"])  

                ws.cell(row=row, column=11, value="---")  # Absent Après-midi
                ws.cell(row=row, column=12, value=p["absence_unique"])  # Absent Après-midi           
                ws.cell(row=row, column=13, value=p["justif_matin"])  # Justificatif Matin
                ws.cell(row=row, column=14, value=p["justif_matin"])  # Justificatif Après-midi
            else:
                ws.cell(row=row, column=5, value=p["matin_entree"])
                ws.cell(row=row, column=6, value=p["matin_sortie"])
                ws.cell(row=row, column=7, value=p["matin_retard"])
                ws.cell(row=row, column=8, value=p["matin_absent"])
                ws.cell(row=row, column=9, value=p["soir_entree"])
                ws.cell(row=row, column=10, value=p["soir_sortie"])
                ws.cell(row=row, column=11, value=p["soir_retard"])
                ws.cell(row=row, column=12, value=p["soir_absent"])
                ws.cell(row=row, column=13, value=p["justif_matin"])
                ws.cell(row=row, column=14, value=p["justif_soir"])

            # Style des cellules de données
            for c in range(1, 15):
                cell = ws.cell(row=row, column=c)
                cell.border = border

                # Alignement horizontal et vertical
                if c == 3:  # Colonne NOM ET PRENOMS
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Wrap text pour les justificatifs longs
                if c >= 13:
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

            row += 1
            idx += 1

    # --- AJUSTEMENT DES LARGEURS ---
    ws.column_dimensions["A"].width = 12  # Date
    ws.column_dimensions["B"].width = 5  # N°
    ws.column_dimensions["C"].width = 52  # Nom
    ws.column_dimensions["D"].width = 10  # IM
    ws.column_dimensions["M"].width = 40  # Justif Matin
    ws.column_dimensions["N"].width = 40  # Justif A-M

    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    os.makedirs("excel_output", exist_ok=True)
    app.run(debug=True)
